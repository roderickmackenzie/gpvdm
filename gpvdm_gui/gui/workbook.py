# 
#   General-purpose Photovoltaic Device Model - a drift diffusion base/Shockley-Read-Hall
#   model for 1st, 2nd and 3rd generation solar cells.
#   Copyright (C) 2012-2017 Roderick C. I. MacKenzie r.c.i.mackenzie at googlemail.com
#
#   https://www.gpvdm.com
#   Room B86 Coates, University Park, Nottingham, NG7 2RD, UK
#
#   This program is free software; you can redistribute it and/or modify
#   it under the terms of the GNU General Public License v2.0, as published by
#   the Free Software Foundation.
#
#   This program is distributed in the hope that it will be useful,
#   but WITHOUT ANY WARRANTY; without even the implied warranty of
#   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#   GNU General Public License for more details.
#
#   You should have received a copy of the GNU General Public License along
#   with this program; if not, write to the Free Software Foundation, Inc.,
#   51 Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.
#
# 

## @package workbook
#  Generate an xls workbook from the simulation
#


work_book_enabled=False
try:
	from openpyxl import Workbook
	from openpyxl.chart import Reference, Series, ScatterChart
	from openpyxl.compat import range
	work_book_enabled=True
except:
	work_book_enabled=False

import glob
import os
from plot_io import plot_load_info
from dat_file import dat_file

from inp import inp_load_file
from inp import inp_get_next_token_array
from token_lib import tokens

from epitaxy import epitaxy_get_dos_files
from epitaxy import epitaxy_get_layers
from epitaxy import epitaxy_get_dos_file
from epitaxy import epitaxy_get_name

def title_truncate(title):
	ret=title

	if len(ret)>30:
		ret=ret[:30]

	return ret

def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False
	
def workbook_from_inp(ws,my_row,filename,title=""):
	lines=[]

	lines=inp_load_file(filename)
	if lines==False:
		return my_row

	if title!="":
		ws.cell(column=1, row=my_row, value=title)
		my_row=my_row+1

	pos=0
	my_token_lib=tokens()

	while (1):
		ret,pos=inp_get_next_token_array(lines,pos)

		token=ret[0]
		if token=="#ver":
			break

		if token=="#end":
			break

		if token.startswith("#"):
			show=False
			units="Units"

			value=ret[1]

			result=my_token_lib.find(token)
			if result!=False:
				units=result.units
				text_info=result.info
				show=True

			if show == True and is_number(value):
				ws.cell(column=1, row=my_row, value=text_info)
				ws.cell(column=2, row=my_row, value=float(value))
				my_row=my_row+1
				
	return my_row

def gen_workbook(input_file_or_dir,output_file):
	if work_book_enabled==False:
		print("python3-openpyxl not found")
		return
	
	wb = Workbook()
	if os.path.isfile(input_file_or_dir):
		files=[input_file_or_dir]
	if os.path.isdir(input_file_or_dir):
		files=glob.glob(os.path.join(input_file_or_dir,"*.dat"))
	else:
		return

	ws = wb.active
	pos=1
	for i in range(0,epitaxy_get_layers()):
		dos_layer=epitaxy_get_dos_file(i)
		if dos_layer.startswith("dos")==True:
			pos=workbook_from_inp(ws,pos,dos_layer+".inp",title=epitaxy_get_name(i))

	
	for my_file in files:
		#print("about to save1",my_file)
		#print(my_file)
		data=dat_file()
		if data.load(my_file,guess=False)==True:
			x=[]
			y=[]
			z=[]
			if data.load(my_file)==True:
				#print("read",my_file)
				ws = wb.create_sheet(title=title_truncate(os.path.basename(my_file)))
				ws.cell(column=1, row=1, value=data.title)
				ws.cell(column=1, row=2, value=data.x_label+" ("+data.x_units+") ")
				ws.cell(column=2, row=2, value=data.data_label+" ("+data.data_units+") ")
		
				for i in range(0,data.y_len):
					ws.cell(column=1, row=i+3, value=data.y_scale[i])
					ws.cell(column=2, row=i+3, value=data.data[0][0][i])

				c1 = ScatterChart()
				c1.title = data.title
				c1.style = 13
				c1.height=20
				c1.width=20
				c1.y_axis.title = data.data_label+" ("+data.data_units+") "
				c1.x_axis.title = data.x_label+" ("+data.x_units+") "

				xdata = Reference(ws, min_col=1, min_row=3, max_row=3+data.y_len)
				ydata = Reference(ws, min_col=2, min_row=3, max_row=3+data.y_len)

				series = Series(ydata,xdata,  title_from_data=True)
				c1.series.append(series)
				ws.add_chart(c1, "G4")
	#print("about to save1")
	try:
		wb.save(filename = output_file)
	except:
		return False

	return True
	#print("about to save0")
