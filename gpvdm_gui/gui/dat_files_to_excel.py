# 
#   General-purpose Photovoltaic Device Model - a drift diffusion base/Shockley-Read-Hall
#   model for 1st, 2nd and 3rd generation solar cells.
#   Copyright (C) 2008-2022 Roderick C. I. MacKenzie r.c.i.mackenzie at googlemail.com
#   
#   https://www.gpvdm.com
#   
#   This program is free software; you can redistribute it and/or modify
#   it under the terms of the GNU General Public License v2.0, as published by
#   the Free Software Foundation.
#   
#   This program is distributed in the hope that it will be useful,
#   but WITHOUT ANY WARRANTY; without even the implied warranty of
#   MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#   GNU General Public License for more details.
#   
#   You should have received a copy of the GNU General Public License along
#   with this program; if not, write to the Free Software Foundation, Inc.,
#   51 Franklin Street, Fifth Floor, Boston, MA 02110-1301 USA.
#   

## @package dat_files_to_excel
#  Load and dump a dat file into a dat class
#

import os
import shutil
import re
from str2bool import str2bool
from util_text import is_number

def dat_files_to_excel(output_file,dat_files):
	try:
		from openpyxl import Workbook
		from openpyxl.chart import Reference, Series, ScatterChart, LineChart
		#from openpyxl.compat import range
	except:
		print("openpyxl not found")
		pass

	if output_file.endswith(".xlsx")==False:
		output_file=output_file+".xlsx"

	wb = Workbook()
	ws_data = wb.active
	#ws_data = wb.create_sheet(title="data")
	data_set=0

	#chart
	c1 = LineChart()
	c1.title = dat_files[0].title
	#c1.style = 13
	c1.height=20
	c1.width=20

	for data in dat_files:
		start_col=data_set*2+1
		ws_data.cell(column=start_col, row=1, value=data.y_label+" ("+data.y_units+") ")
		ws_data.cell(column=start_col+1, row=1, value=data.data_label+" ("+data.data_units+") ")

		row_pos=0
		for i in range(0,data.y_len):
			row_pos=i+2
			ws_data.cell(column=start_col, row=row_pos, value=data.y_scale[i])
			ws_data.cell(column=start_col+1, row=row_pos, value=data.data[0][0][i])

		data = Reference(ws_data, min_col=start_col, max_col=start_col+1, min_row=1, max_row=row_pos)
		c1.add_data(data, titles_from_data=True)

		data_set=data_set+1


	c1.y_axis.title = dat_files[0].data_label+" ("+dat_files[0].data_units+") "
	c1.x_axis.title = dat_files[0].y_label+" ("+dat_files[0].y_units+") "


	ws_data.add_chart(c1, "G4")


	#print("about to save1")
	try:
		wb.save(filename = output_file)
	except:
		return False

	return
	max=data[0].y_len
	for d in data:
		if d.y_len>max:
			max=d.y_len
	out=[]
	line=""
	for i in range(0,len(data)):
		line=line+data[i].key_text+" "+data[i].y_label+","+data[i].key_text+" "+data[i].data_label+","

	line=line[:-1]
	out.append(line)

	for i in range(0,max):
		line=""
		for ii in range(0,len(data)):
			y_text=""
			data_text=""
			if i<data[ii].y_len:
				y_text=str('{:.8e}'.format(float(data[ii].y_scale[i])))
				data_text=str('{:.8e}'.format(float(data[ii].data[0][0][i])))
			line=line+y_text+","+data_text+","
		line=line[:-1]
		out.append(line)

	inp_save_lines_to_file(file_name,out)


